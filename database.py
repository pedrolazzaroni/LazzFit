import sqlite3
from datetime import datetime
import csv
import os
import time
import sys
import traceback
import shutil  # Para operações de cópia de arquivos

# Verificar a disponibilidade do módulo openpyxl
EXCEL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
    print("✅ Módulo openpyxl encontrado - Exportação para Excel disponível.")
except ImportError:
    print("⚠️ Módulo 'openpyxl' não encontrado. A exportação para Excel não estará disponível.")
    print("Para instalar o módulo, execute: pip install openpyxl")

class DatabaseManager:
    def __init__(self, db_name):
        # Garantir que o caminho do banco de dados é absoluto e único
        if not os.path.isabs(db_name):
            app_dir = os.path.dirname(os.path.abspath(__file__))
            data_dir = os.path.join(app_dir, "data")
            if not os.path.exists(data_dir):
                os.makedirs(data_dir)
            self.db_name = os.path.join(data_dir, os.path.basename(db_name))
        else:
            self.db_name = db_name
            
        print(f"🔍 DatabaseManager usando banco de dados em: {self.db_name}")
        
        # Verificar se o diretório do banco existe
        db_dir = os.path.dirname(self.db_name)
        if not os.path.exists(db_dir):
            try:
                os.makedirs(db_dir)
                print(f"✓ Diretório do banco de dados criado: {db_dir}")
            except Exception as e:
                print(f"❌ Erro ao criar diretório do banco de dados: {e}")
        
        self.conn = None
        self.cursor = None
        
        # Tipos de treino pré-definidos
        self.workout_types = [
            "Corrida de Rua",
            "Trail Running",
            "Corrida na Esteira",
            "Intervalado",
            "Long Run",
            "Recuperação",
            "Fartlek",
            "Corrida de Montanha",
            "Outro"
        ]
        
        self.transaction_active = False  # Controle de transação ativa
        
        # Atributo para indicar disponibilidade de Excel
        self.EXCEL_AVAILABLE = EXCEL_AVAILABLE
        
    def connect(self):
        """Conecta ao banco de dados (sem iniciar transação)"""
        try:
            if self.conn:
                # Verificar se a conexão ainda está ativa
                try:
                    self.conn.execute("SELECT 1")
                    return True
                except sqlite3.Error:
                    # Conexão não está mais válida, fechamos e tentamos novamente
                    try:
                        self.conn.close()
                    except:
                        pass
                    self.conn = None
                    self.cursor = None
            
            # Verificar se o diretório do banco existe, se não, criar
            db_dir = os.path.dirname(self.db_name)
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir)
                
            # Tenta conectar com timeout para evitar problemas de lock
            print(f"🔌 Tentando conectar ao banco de dados: {self.db_name}")
            
            # IMPORTANTE: Usar DEFERRED para controle explícito de transações
            self.conn = sqlite3.connect(self.db_name, timeout=30, isolation_level="DEFERRED")
            
            # Configurações para melhorar a persistência
            self.conn.execute("PRAGMA foreign_keys = ON")
            # Configuração mais robusta de journal
            self.conn.execute("PRAGMA journal_mode = DELETE")
            # Modo FULL para máxima segurança de dados
            self.conn.execute("PRAGMA synchronous = NORMAL")  # NORMAL é bom equilíbrio entre segurança e performance
            
            self.cursor = self.conn.cursor()
            return True
            
        except sqlite3.Error as e:
            print(f"❌ Erro de conexão ao banco: {e}")
            return False
    
    def begin_transaction(self):
        """Inicia uma transação explicitamente"""
        if self.conn and not self.transaction_active:
            try:
                self.conn.execute("BEGIN TRANSACTION")
                self.transaction_active = True
                return True
            except sqlite3.Error as e:
                print(f"❌ Erro ao iniciar transação: {e}")
                return False
        return self.transaction_active  # Já estava ativa
    
    def commit(self):
        """Faz commit da transação se houver uma ativa"""
        if self.conn and self.transaction_active:
            try:
                self.conn.commit()
                self.transaction_active = False
                return True
            except sqlite3.Error as e:
                print(f"❌ Erro ao fazer commit: {e}")
                return False
        return True  # Nada para fazer commit
    
    def rollback(self):
        """Desfaz a transação se houver uma ativa"""
        if self.conn and self.transaction_active:
            try:
                self.conn.rollback()
                self.transaction_active = False
                return True
            except sqlite3.Error as e:
                print(f"❌ Erro ao fazer rollback: {e}")
                return False
        return True  # Nada para fazer rollback
    
    def disconnect(self):
        """Desconecta do banco de dados, fazendo commit se necessário"""
        if self.conn:
            try:
                # Commit se houver uma transação ativa
                if self.transaction_active:
                    self.commit()
                
                self.conn.close()
                print("✓ Banco de dados fechado com sucesso")
            except sqlite3.Error as e:
                print(f"❌ Erro ao fechar conexão com banco de dados: {e}")
            finally:
                self.conn = None
                self.cursor = None
                self.transaction_active = False
    
    def setup(self):
        """Configura o banco de dados com tratamento de erros melhorado"""
        if not self.connect():
            print("❌ ERRO: Não foi possível conectar ao banco de dados para setup.")
            return False
            
        try:
            # Iniciar uma transação para todas as operações de setup
            self.begin_transaction()
            
            # Verificar se a tabela existe
            self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='runs'")
            table_exists = self.cursor.fetchone()
                
            # Criar a tabela se não existir
            if not table_exists:
                print("Tabela 'runs' não existe. Criando nova tabela...")
                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT,
                    distance REAL,
                    duration INTEGER,
                    avg_pace TEXT,
                    avg_bpm INTEGER,
                    max_bpm INTEGER,
                    elevation_gain INTEGER,
                    calories INTEGER,
                    workout_type TEXT,
                    notes TEXT
                )
                ''')
                print("✓ Tabela 'runs' criada com sucesso")
            else:
                print("✓ Tabela 'runs' já existe")
                # Verificar colunas faltantes
                self.check_and_add_missing_columns()
            
            # Configurar tabelas para planilhas de treino
            self._setup_training_plan_tables()
            
            # Confirmar as alterações
            self.commit()
            return True
            
        except Exception as e:
            print(f"❌ ERRO durante setup do banco de dados: {e}")
            print(traceback.format_exc())
            self.rollback()  # Desfazer alterações em caso de erro
            return False
        finally:
            self.disconnect()
    
    def check_and_add_missing_columns(self):
        """Verifica e adiciona colunas faltantes na tabela runs"""
        try:
            self.cursor.execute("PRAGMA table_info(runs)")
            columns = self.cursor.fetchall()
            column_names = [col[1] for col in columns]
            
            missing_columns = []
            if "avg_bpm" not in column_names:
                missing_columns.append("avg_bpm INTEGER")
            if "max_bpm" not in column_names:
                missing_columns.append("max_bpm INTEGER")
            if "elevation_gain" not in column_names:
                missing_columns.append("elevation_gain INTEGER")
            if "workout_type" not in column_names:
                missing_columns.append("workout_type TEXT")
            
            # Adicionar colunas faltantes
            for col in missing_columns:
                col_name = col.split()[0]
                try:
                    self.cursor.execute(f"ALTER TABLE runs ADD COLUMN {col}")
                    print(f"✓ Coluna '{col_name}' adicionada à tabela 'runs'")
                except sqlite3.OperationalError:
                    # Coluna já existe
                    pass
            
            return True
        except sqlite3.Error as e:
            print(f"❌ Erro ao verificar colunas da tabela: {e}")
            return False
    
    def _setup_training_plan_tables(self):
        """Configura as tabelas necessárias para planilhas de treino"""
        try:
            # Tabela principal de planos de treino
            self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS training_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                goal TEXT,
                duration_weeks INTEGER,
                level TEXT, 
                notes TEXT,
                created_at TEXT,
                updated_at TEXT
            )
            ''')
            
            # Tabela para semanas de treino
            self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS training_weeks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER,
                week_number INTEGER,
                focus TEXT,
                total_distance REAL,
                notes TEXT,
                FOREIGN KEY (plan_id) REFERENCES training_plans (id) ON DELETE CASCADE
            )
            ''')
            
            # Tabela para sessões individuais de treino
            self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS training_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_id INTEGER,
                day_of_week INTEGER, 
                workout_type TEXT,
                distance REAL,
                duration INTEGER,
                intensity TEXT,
                pace_target TEXT,
                hr_zone TEXT,
                details TEXT,
                FOREIGN KEY (week_id) REFERENCES training_weeks (id) ON DELETE CASCADE
            )
            ''')
            
            print("✓ Tabelas de planos de treino configuradas com sucesso")
            return True
        except sqlite3.Error as e:
            print(f"❌ Erro ao configurar tabelas de planos de treino: {e}")
            return False

    def add_run(self, date, distance, duration, avg_bpm, max_bpm, elevation_gain, calories, workout_type, notes=""):
        """Adiciona um novo registro de corrida"""
        if not self.connect():
            print("ERRO: Não foi possível conectar ao banco de dados para adicionar corrida.")
            return None
            
        try:
            # Iniciar transação
            self.begin_transaction()
            
            # Calcula o ritmo médio (pace) em formato min:seg por km
            if distance > 0:
                pace_seconds = (duration * 60) / distance
                minutes = int(pace_seconds // 60)
                seconds = int(pace_seconds % 60)
                avg_pace = f"{minutes}:{seconds:02d}"
            else:
                avg_pace = "0:00"
                
            self.cursor.execute(
                "INSERT INTO runs (date, distance, duration, avg_pace, avg_bpm, max_bpm, elevation_gain, calories, workout_type, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (date, distance, duration, avg_pace, avg_bpm, max_bpm, elevation_gain, calories, workout_type, notes)
            )
            
            # Commit da transação
            self.commit()
            print(f"✓ Adicionada corrida em {date}")
            run_id = self.cursor.lastrowid
            return run_id
        except sqlite3.Error as e:
            print(f"ERRO ao adicionar corrida: {e}")
            self.rollback()
            return None
        finally:
            self.disconnect()
    
    def get_all_runs(self):
        """Retorna todos os registros de corrida"""
        if not self.connect():
            print("ERRO: Não foi possível conectar ao banco de dados para obter corridas.")
            return []
            
        try:
            self.cursor.execute("SELECT * FROM runs ORDER BY date DESC")
            runs = self.cursor.fetchall()
            return runs
        except sqlite3.Error as e:
            print(f"ERRO ao obter corridas: {e}")
            return []
        finally:
            self.disconnect()
    
    def get_run(self, run_id):
        """Retorna um registro de corrida específico pelo ID"""
        if not self.connect():
            return None
            
        try:
            self.cursor.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
            run = self.cursor.fetchone()
            return run
        except sqlite3.Error as e:
            print(f"ERRO ao obter corrida {run_id}: {e}")
            return None
        finally:
            self.disconnect()
    
    def update_run(self, run_id, date, distance, duration, avg_bpm, max_bpm, elevation_gain, calories, workout_type, notes=""):
        """Atualiza um registro de corrida existente"""
        if not self.connect():
            print("ERRO: Não foi possível conectar ao banco de dados para atualizar corrida.")
            return False
            
        try:
            # Iniciar transação
            self.begin_transaction()
            
            # Calcula o ritmo médio (pace) em formato min:seg por km
            if distance > 0:
                pace_seconds = (duration * 60) / distance
                minutes = int(pace_seconds // 60)
                seconds = int(pace_seconds % 60)
                avg_pace = f"{minutes}:{seconds:02d}"
            else:
                avg_pace = "0:00"
                
            self.cursor.execute(
                "UPDATE runs SET date=?, distance=?, duration=?, avg_pace=?, avg_bpm=?, max_bpm=?, elevation_gain=?, calories=?, workout_type=?, notes=? WHERE id=?",
                (date, distance, duration, avg_pace, avg_bpm, max_bpm, elevation_gain, calories, workout_type, notes, run_id)
            )
            
            # Commit da transação
            self.commit()
            print(f"✓ Atualizada corrida ID {run_id}")
            return True
        except sqlite3.Error as e:
            print(f"ERRO ao atualizar corrida {run_id}: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
    
    def delete_run(self, run_id):
        """Exclui um registro de corrida pelo ID"""
        if not self.connect():
            print("ERRO: Não foi possível conectar ao banco de dados para excluir corrida.")
            return False
            
        try:
            # Iniciar transação
            self.begin_transaction()
            
            self.cursor.execute("DELETE FROM runs WHERE id=?", (run_id,))
            
            # Commit da transação
            self.commit()
            print(f"✓ Excluída corrida ID {run_id}")
            return True
        except sqlite3.Error as e:
            print(f"ERRO ao excluir corrida {run_id}: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
        
    def export_runs_to_csv(self, file_path, run_ids=None):
        """Exporta os treinos para um arquivo CSV"""
        if not self.connect():
            return False
        
        try:
            if run_ids:
                # Converte lista de IDs para formato adequado para IN do SQL
                id_placeholders = ','.join(['?' for _ in run_ids])
                query = f"SELECT * FROM runs WHERE id IN ({id_placeholders}) ORDER BY date DESC"
                self.cursor.execute(query, run_ids)
            else:
                self.cursor.execute("SELECT * FROM runs ORDER BY date DESC")
                
            runs = self.cursor.fetchall()
            
            if not runs:
                return False
                
            # Nome das colunas para o cabeçalho do CSV
            headers = [
                'ID', 'Data', 'Distância (km)', 'Duração (min)', 'Ritmo (min/km)', 
                'BPM Médio', 'BPM Máximo', 'Ganho de Elevação (m)', 'Calorias', 
                'Tipo de Treino', 'Notas'
            ]
            
            with open(file_path, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerows(runs)
            return True
        except Exception as e:
            print(f"Erro ao exportar dados para CSV: {e}")
            return False
        finally:
            self.disconnect()
            
    def export_runs_to_xlsx(self, file_path, run_ids=None):
        """Exporta os treinos para um arquivo Excel"""
        # Verificar novamente se o openpyxl está disponível
        excel_available = False
        try:
            import openpyxl
            excel_available = True
        except ImportError:
            excel_available = False
            
        if not excel_available:
            print("⚠️ A exportação para Excel não está disponível porque o módulo 'openpyxl' não está instalado.")
            return False
            
        if not self.connect():
            return False
        
        try:
            if run_ids:
                # Converte lista de IDs para formato adequado para IN do SQL
                id_placeholders = ','.join(['?' for _ in run_ids])
                query = f"SELECT * FROM runs WHERE id IN ({id_placeholders}) ORDER BY date DESC"
                self.cursor.execute(query, run_ids)
            else:
                self.cursor.execute("SELECT * FROM runs ORDER BY date DESC")
                
            runs = self.cursor.fetchall()
            
            if not runs:
                return False
                
            # Nome das colunas para o cabeçalho 
            headers = [
                'ID', 'Data', 'Distância (km)', 'Duração (min)', 'Ritmo (min/km)', 
                'BPM Médio', 'BPM Máximo', 'Ganho de Elevação (m)', 'Calorias', 
                'Tipo de Treino', 'Notas'
            ]
            
            # Criar um novo workbook e selecionar a planilha ativa
            wb = Workbook()
            ws = wb.active
            ws.title = "Treinos de Corrida"
            
            # Estilos
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="FF6700", end_color="FF6700", fill_type="solid")
            centered_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(border_style="thin", color="DDDDDD"),
                right=Side(border_style="thin", color="DDDDDD"),
                top=Side(border_style="thin", color="DDDDDD"),
                bottom=Side(border_style="thin", color="DDDDDD")
            )

            # Adicionar cabeçalho
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                cell.border = border
                
            # Ajustar a largura das colunas baseado nos cabeçalhos
            for col_num, header in enumerate(headers, 1):
                column_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[column_letter].width = max(len(header) * 1.2, 12)

            # Adicionar dados
            for row_num, run in enumerate(runs, 2):
                for col_num, value in enumerate(run, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Tratamento especial para alguns valores
                    if col_num == 3:  # Distância
                        cell.value = round(value, 2) if value else 0
                        cell.number_format = '0.00'
                    elif col_num in (6, 7, 8):  # BPM e Elevação
                        cell.value = value if value is not None else "N/A"
                    else:
                        cell.value = value if value is not None else ""
                    
                    cell.alignment = Alignment(vertical='center')
                    cell.border = border
                
                # Aplicar formatação alternada para linhas
                row_color = "F5F5F5" if row_num % 2 == 0 else "FFFFFF"
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_num).fill = PatternFill(
                        start_color=row_color, end_color=row_color, fill_type="solid"
                    )
            
            # Salvar o workbook
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Erro ao exportar dados para Excel: {e}")
            return False
        finally:
            self.disconnect()
    
    def get_workout_types(self):
        """Retorna a lista de tipos de treino disponíveis"""
        return self.workout_types
    
    def ensure_connection_closed(self):
        """Garantir que a conexão esteja fechada"""
        if self.conn:
            try:
                # Se houver uma transação ativa, fazer commit
                if self.transaction_active:
                    self.commit()
                self.conn.close()
                print("✓ Conexão com banco fechada através de ensure_connection_closed")
            except Exception as e:
                print(f"Erro ao fechar conexão: {e}")
            finally:
                self.conn = None
                self.cursor = None
                self.transaction_active = False
    
    def __del__(self):
        """Destrutor para garantir que a conexão seja fechada"""
        self.ensure_connection_closed()
    
    # Métodos para gerenciar planos de treino
    def create_training_plan(self, name, goal, duration_weeks, level, notes=""):
        """Cria um novo plano de treinamento"""
        if not self.connect():
            return None
            
        try:
            self.begin_transaction()
            
            # Data de criação atual
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            self.cursor.execute(
                "INSERT INTO training_plans (name, goal, duration_weeks, level, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (name, goal, duration_weeks, level, notes, created_at, created_at)
            )
            
            plan_id = self.cursor.lastrowid
            
            # Criar semanas vazias para o plano
            for week_num in range(1, duration_weeks + 1):
                self.cursor.execute(
                    "INSERT INTO training_weeks (plan_id, week_number, focus, total_distance, notes) VALUES (?, ?, ?, ?, ?)",
                    (plan_id, week_num, f"Semana {week_num}", 0.0, "")
                )
                
                week_id = self.cursor.lastrowid
                
                # Criar sessões vazias para cada dia da semana
                for day in range(1, 8):  # 1=Segunda, 7=Domingo
                    self.cursor.execute(
                        "INSERT INTO training_sessions (week_id, day_of_week, workout_type, distance, duration, intensity, details) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (week_id, day, "Descanso", 0.0, 0, "Baixa", "")
                    )
            
            self.commit()
            return plan_id
        except sqlite3.Error as e:
            print(f"Erro ao criar plano de treino: {e}")
            self.rollback()
            return None
        finally:
            self.disconnect()
    
    def get_all_training_plans(self):
        """Retorna todos os planos de treinamento"""
        if not self.connect():
            return []
            
        try:
            self.cursor.execute("SELECT * FROM training_plans ORDER BY updated_at DESC")
            plans = self.cursor.fetchall()
            return plans
        except sqlite3.Error as e:
            print(f"Erro ao obter planos de treinamento: {e}")
            return []
        finally:
            self.disconnect()
    
    def get_training_plan(self, plan_id):
        """Retorna um plano de treinamento específico com todas as suas semanas e sessões"""
        if not self.connect():
            return None
            
        try:
            # Obter informações do plano
            self.cursor.execute("SELECT * FROM training_plans WHERE id = ?", (plan_id,))
            plan = self.cursor.fetchone()
            
            if not plan:
                return None
                
            # Obter todas as semanas
            self.cursor.execute("SELECT * FROM training_weeks WHERE plan_id = ? ORDER BY week_number", (plan_id,))
            weeks = self.cursor.fetchall()
            
            # Obter todas as sessões para cada semana
            plan_data = {
                "id": plan[0],
                "name": plan[1],
                "goal": plan[2],
                "duration_weeks": plan[3],
                "level": plan[4],
                "notes": plan[5],
                "created_at": plan[6],
                "updated_at": plan[7],
                "weeks": []
            }
            
            for week in weeks:
                week_id = week[0]
                self.cursor.execute("SELECT * FROM training_sessions WHERE week_id = ? ORDER BY day_of_week", (week_id,))
                sessions = self.cursor.fetchall()
                
                week_data = {
                    "id": week[0],
                    "week_number": week[2],
                    "focus": week[3],
                    "total_distance": week[4],
                    "notes": week[5],
                    "sessions": []
                }
                
                for session in sessions:
                    session_data = {
                        "id": session[0],
                        "day_of_week": session[2],
                        "workout_type": session[3],
                        "distance": session[4],
                        "duration": session[5],
                        "intensity": session[6],
                        "pace_target": session[7],
                        "hr_zone": session[8],
                        "details": session[9]
                    }
                    week_data["sessions"].append(session_data)
                
                plan_data["weeks"].append(week_data)
            
            return plan_data
        except sqlite3.Error as e:
            print(f"Erro ao obter plano de treinamento: {e}")
            return None
        finally:
            self.disconnect()
    
    def update_training_plan(self, plan_id, name, goal, duration_weeks, level, notes):
        """Atualiza um plano de treinamento existente"""
        if not self.connect():
            return False
            
        try:
            self.begin_transaction()
            
            updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            self.cursor.execute(
                "UPDATE training_plans SET name=?, goal=?, duration_weeks=?, level=?, notes=?, updated_at=? WHERE id=?",
                (name, goal, duration_weeks, level, notes, updated_at, plan_id)
            )
            
            self.commit()
            return True
        except sqlite3.Error as e:
            print(f"Erro ao atualizar plano de treinamento: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
    
    def update_training_week(self, week_id, focus, total_distance, notes):
        """Atualiza uma semana de treinamento"""
        if not self.connect():
            return False
            
        try:
            self.begin_transaction()
            
            self.cursor.execute(
                "UPDATE training_weeks SET focus=?, total_distance=?, notes=? WHERE id=?",
                (focus, total_distance, notes, week_id)
            )
            
            # Atualizar o timestamp no plano principal
            self.cursor.execute(
                """
                UPDATE training_plans SET updated_at=?
                WHERE id=(SELECT plan_id FROM training_weeks WHERE id=?)
                """,
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), week_id)
            )
            
            self.commit()
            return True
        except sqlite3.Error as e:
            print(f"Erro ao atualizar semana de treinamento: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
    
    def update_training_session(self, session_id, workout_type, distance, duration, intensity, pace_target, hr_zone, details):
        """Atualiza uma sessão de treinamento"""
        if not self.connect():
            return False
            
        try:
            self.begin_transaction()
            
            self.cursor.execute(
                "UPDATE training_sessions SET workout_type=?, distance=?, duration=?, intensity=?, pace_target=?, hr_zone=?, details=? WHERE id=?",
                (workout_type, distance, duration, intensity, pace_target, hr_zone, details, session_id)
            )
            
            # Atualizar o timestamp no plano principal
            self.cursor.execute(
                """
                UPDATE training_plans SET updated_at=?
                WHERE id=(SELECT plan_id FROM training_weeks WHERE id=(SELECT week_id FROM training_sessions WHERE id=?))
                """,
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session_id)
            )
            
            self.commit()
            return True
        except sqlite3.Error as e:
            print(f"Erro ao atualizar sessão de treinamento: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
    
    def delete_training_plan(self, plan_id):
        """Exclui um plano de treinamento"""
        if not self.connect():
            return False
            
        try:
            self.begin_transaction()
            
            # As exclusões em cascata serão tratadas pelos FKs
            self.cursor.execute("DELETE FROM training_plans WHERE id=?", (plan_id,))
            
            self.commit()
            return True
        except sqlite3.Error as e:
            print(f"Erro ao excluir plano de treinamento: {e}")
            self.rollback()
            return False
        finally:
            self.disconnect()
    
    def export_training_plan_to_xlsx(self, plan_id, file_path):
        """Exporta um plano de treinamento para arquivo Excel"""
        # Verificar se o openpyxl está disponível
        excel_available = False
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            excel_available = True
        except ImportError:
            return False
            
        if not excel_available:
            return False
            
        if not self.connect():
            return False
            
        try:
            # Obter todas as informações do plano
            plan_data = self.get_training_plan(plan_id)
            
            if not plan_data:
                return False
                
            # Criar workbook
            wb = openpyxl.Workbook()
            
            # Configurar estilos
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="FF6700", end_color="FF6700", fill_type="solid")
            subheader_font = Font(name='Arial', size=11, bold=True)
            subheader_fill = PatternFill(start_color="FFC299", end_color="FFC299", fill_type="solid")
            title_font = Font(name='Arial', size=16, bold=True)
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Planilha de informações gerais
            ws_info = wb.active
            ws_info.title = "Informações do Plano"
            
            ws_info['A1'] = "PLANO DE TREINO DE CORRIDA"
            ws_info['A1'].font = title_font
            ws_info.merge_cells('A1:F1')
            ws_info['A1'].alignment = center_align
            
            # Informações básicas do plano
            ws_info['A3'] = "Nome do Plano:"
            ws_info['B3'] = plan_data["name"]
            ws_info['A4'] = "Objetivo:"
            ws_info['B4'] = plan_data["goal"]
            ws_info['A5'] = "Duração (semanas):"
            ws_info['B5'] = plan_data["duration_weeks"]
            ws_info['A6'] = "Nível:"
            ws_info['B6'] = plan_data["level"]
            ws_info['A7'] = "Notas:"
            ws_info['B7'] = plan_data["notes"]
            
            # Planilha para cada semana
            for week in plan_data["weeks"]:
                week_num = week["week_number"]
                ws = wb.create_sheet(f"Semana {week_num}")
                
                # Cabeçalho da semana
                ws['A1'] = f"SEMANA {week_num} - {week['focus']}"
                ws['A1'].font = title_font
                ws.merge_cells('A1:H1')
                ws['A1'].alignment = center_align
                
                # Cabeçalhos de dias da semana
                day_names = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
                for i, day in enumerate(day_names):
                    col = chr(66 + i)  # B=Segunda, C=Terça, etc.
                    ws[f'{col}3'] = day
                    ws[f'{col}3'].font = header_font
                    ws[f'{col}3'].fill = header_fill
                    ws[f'{col}3'].border = border
                    ws[f'{col}3'].alignment = center_align
                
                # Cabeçalho de atributos
                attributes = ["Tipo de Treino", "Distância (km)", "Duração (min)", "Intensidade", 
                              "Pace Alvo", "Zona de FC", "Detalhes"]
                for i, attr in enumerate(attributes):
                    ws[f'A{4+i}'] = attr
                    ws[f'A{4+i}'].font = subheader_font
                    ws[f'A{4+i}'].border = border
                
                # Preencher dados das sessões
                for session in week["sessions"]:
                    day_idx = session["day_of_week"]
                    col = chr(65 + day_idx)  # B=Segunda (1), C=Terça (2), etc.
                    
                    # Tipo de treino
                    ws[f'{col}4'] = session["workout_type"]
                    # Distância
                    ws[f'{col}5'] = session["distance"] if session["distance"] > 0 else "-"
                    # Duração
                    ws[f'{col}6'] = session["duration"] if session["duration"] > 0 else "-"
                    # Intensidade
                    ws[f'{col}7'] = session["intensity"]
                    # Pace alvo
                    ws[f'{col}8'] = session["pace_target"] if session["pace_target"] else "-"
                    # Zona FC
                    ws[f'{col}9'] = session["hr_zone"] if session["hr_zone"] else "-"
                    # Detalhes
                    ws[f'{col}10'] = session["details"] if session["details"] else "-"
                    
                    # Adicionar bordas
                    for row in range(4, 11):
                        ws[f'{col}{row}'].border = border
                
                # Ajustar largura das colunas
                for col in range(ord('A'), ord('I')):
                    ws.column_dimensions[chr(col)].width = 15
                
                # Notas da semana
                ws['A12'] = "Notas da Semana:"
                ws['A12'].font = subheader_font
                ws['B12'] = week["notes"]
                ws.merge_cells('B12:H12')
                
                # Distância total da semana
                ws['A13'] = "Distância Total:"
                ws['A13'].font = subheader_font
                ws['B13'] = f"{week['total_distance']} km"
            
            # Salvar o workbook
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Erro ao exportar plano de treinamento para Excel: {e}")
            traceback.print_exc()
            return False
        finally:
            self.disconnect()
